import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook

def processar_planilha():
    caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if not caminho:
        return

    def wrap_cell(cell):
        if not cell:
            return ""
        partes = str(cell).splitlines()
        return " ".join(f"<p>{p.strip()}</p>" for p in partes if p.strip())

    wb_in = load_workbook(caminho)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active

    for i, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row, start=1):
            ws_out.cell(row=i, column=j, value=wrap_cell(cell))

    novo_caminho = caminho.replace(".xlsx", "_tratado.xlsx")
    wb_out.save(novo_caminho)

    messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{novo_caminho}")

janela = tk.Tk()
janela.title("Processador de Planilhas")

botao = tk.Button(janela, text="Selecionar e Processar Planilha", command=processar_planilha)
botao.pack(padx=20, pady=20)

janela.mainloop()