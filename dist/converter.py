from openpyxl import load_workbook, Workbook

# Arquivos
input_file = "IA.xlsx"
output_file = "IA_tratado.xlsx"

# Função que adiciona <p> em cada valor
def wrap_cell(cell):
    if not cell:
        return ""
    
    # separa por quebra de linha (caso tenha vários itens)
    partes = str(cell).splitlines()
    
    # junta tudo com <p> na mesma linha (igual você pediu)
    return " ".join(f"<p>{p.strip()}</p>" for p in partes if p.strip())

# Abrir arquivo original
wb_in = load_workbook(input_file)
ws_in = wb_in.active

# Criar novo arquivo
wb_out = Workbook()
ws_out = wb_out.active

# Percorrer células
for i, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
    for j, cell in enumerate(row, start=1):
        ws_out.cell(row=i, column=j, value=wrap_cell(cell))

# Salvar resultado
wb_out.save(output_file)

print("Arquivo gerado com sucesso!")